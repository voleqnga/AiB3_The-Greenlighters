from flask import Flask, request, jsonify
from flask_cors import CORS
import json
import os
import re
import io
import base64
import joblib
import numpy as np
import urllib.request
import urllib.error
from datetime import datetime
import uuid

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

app = Flask(__name__)
CORS(app)

# Thư mục project (cùng cấp với server.py) — luôn ghi vào fixdemo99/output
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')
os.makedirs(OUTPUT_DIR, exist_ok=True)

MODEL_DIR = os.path.join(BASE_DIR, 'model')
PREPROCESS_DIR = os.path.join(BASE_DIR, 'preprocess')
for _d in (MODEL_DIR, PREPROCESS_DIR):
    os.makedirs(_d, exist_ok=True)

HR_POOL_DIR = os.path.join(OUTPUT_DIR, 'hr_pool')
os.makedirs(HR_POOL_DIR, exist_ok=True)
# Demo: chỉ một JD + một pool ứng viên (không dùng jdId)
HR_POOL_FILE = os.path.join(HR_POOL_DIR, 'candidates.json')

CV_OUTPUT_DIR = os.path.join(OUTPUT_DIR, 'CV')
JD_OUTPUT_DIR = os.path.join(OUTPUT_DIR, 'JD')
SURVEY_OUTPUT_DIR = os.path.join(OUTPUT_DIR, 'Survey')
for _d in (CV_OUTPUT_DIR, JD_OUTPUT_DIR, SURVEY_OUTPUT_DIR):
    os.makedirs(_d, exist_ok=True)

SURVEY_EXCEL_FILE = os.path.join(SURVEY_OUTPUT_DIR, 'survey_responses.xlsx')

JD_CURRENT_FILE = os.path.join(JD_OUTPUT_DIR, 'JD_current.json')
# Snapshot phân tích AI + file CV JSON — cùng thư mục output/CV (trước đây là output/Analysis)
ANALYSIS_DIR = CV_OUTPUT_DIR


def _safe_analysis_filename(name):
    """Tên file an toàn cho Analysis_<name>.json (giữ Unicode chữ cái tiếng Việt)."""
    raw = (name or '').strip() or 'Ung_vien'
    safe = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', raw)
    safe = safe.strip(' .') or 'Ung_vien'
    if len(safe) > 120:
        safe = safe[:120]
    return safe


def _safe_cv_name_stem(name):
    """Cùng logic với upload_cv → CV_<stem>.json: space → _, ký tự không phải word → _, tối đa 60 ký tự."""
    stem = (name or 'Unknown').replace(' ', '_')
    stem = re.sub(r'[^\w\-]', '_', stem, flags=re.UNICODE)[:60]
    return stem.strip('_') or 'Unknown'


def _survey_column_order(keys):
    """Thứ tự cột ổn định cho Excel; JSON_File luôn cuối."""
    priority = [
        'Timestamp', 'CV_ID', 'Candidate_Name', 'candidate_name',
        'Overall_Match_Score',
    ]
    q_pairs = []
    for i in range(1, 13):
        q_pairs.extend([f'Q{i}_Rating', f'Q{i}_Comment'])
    priority.extend(q_pairs)
    priority.append('Additional_Feedback')
    seen = set()
    ordered = []
    for k in priority:
        if k in keys and k not in seen:
            ordered.append(k)
            seen.add(k)
    rest = sorted(k for k in keys if k not in seen and k != 'JSON_File')
    ordered.extend(rest)
    return ordered


def _survey_column_width(header_name):
    """Độ rộng cột (Excel units) theo loại cột — dễ đọc, comment rộng hơn."""
    h = header_name or ''
    if h == 'JSON_File':
        return 36
    if h == 'Timestamp':
        return 24
    if h in ('CV_ID', 'Overall_Match_Score'):
        return 16
    if h in ('Candidate_Name', 'candidate_name'):
        return 28
    if 'Comment' in h or h == 'Additional_Feedback':
        return 44
    if 'Rating' in h:
        return 12
    return 20


def _style_survey_worksheet(ws, cols):
    """Định dạng bảng: chữ lớn, header nổi bật, viền, freeze, xen kẽ dòng."""
    try:
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    if not cols:
        return

    max_row = ws.max_row or 1
    max_col = len(cols)

    thin = Side(style='thin', color='CBD5E1')
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill(fill_type='solid', fgColor='1E40AF')
    header_font = Font(bold=True, size=13, color='FFFFFF', name='Calibri')
    body_font = Font(size=12, name='Calibri')
    alt_fill = PatternFill(fill_type='solid', fgColor='F1F5F9')
    white_fill = PatternFill(fill_type='solid', fgColor='FFFFFF')

    for idx, name in enumerate(cols, start=1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = _survey_column_width(name)

    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = grid

    ws.row_dimensions[1].height = 26
    ws.freeze_panes = 'A2'

    for r in range(2, max_row + 1):
        row_fill = alt_fill if r % 2 == 0 else white_fill
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = body_font
            cell.border = grid
            cell.fill = row_fill
            col_name = cols[c - 1]
            wrap = (
                'Comment' in col_name
                or col_name == 'Additional_Feedback'
                or col_name == 'Timestamp'
            )
            if 'Rating' in col_name or col_name == 'Overall_Match_Score':
                h_align = 'center'
            elif col_name == 'JSON_File':
                h_align = 'left'
            else:
                h_align = 'left'
            cell.alignment = Alignment(
                horizontal=h_align,
                vertical='top',
                wrap_text=wrap,
            )
        ws.row_dimensions[r].height = 22


def _list_survey_json_files():
    """Các file survey_*.json trong thư mục Survey (không tính survey_responses.*)."""
    out = []
    if not os.path.isdir(SURVEY_OUTPUT_DIR):
        return out
    for name in sorted(os.listdir(SURVEY_OUTPUT_DIR)):
        if not name.endswith('.json'):
            continue
        if name == 'survey_responses.json':
            continue
        if name.startswith('survey_'):
            out.append(os.path.join(SURVEY_OUTPUT_DIR, name))
    return out


def _survey_excel_file_ok(path):
    """True nếu file tồn tại, không rỗng và mở được bằng openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return False
    if not os.path.isfile(path):
        return False
    if os.path.getsize(path) == 0:
        return False
    try:
        wb = load_workbook(path)
        wb.close()
        return True
    except Exception:
        return False


def rebuild_survey_excel_from_json_files():
    """Tạo lại survey_responses.xlsx từ mọi survey_*.json (phục hồi khi file xlsx rỗng/hỏng)."""
    try:
        from openpyxl import Workbook
    except ImportError:
        print('[Survey] openpyxl chua cai — khong tao lai Excel')
        return False

    paths = _list_survey_json_files()
    if not paths:
        return False

    all_rows = []
    all_keys = set()
    for path in paths:
        basename = os.path.basename(path)
        try:
            with open(path, encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            print(f'[Survey] Bo qua {basename}: {e}')
            continue
        if not isinstance(data, dict):
            continue
        row = dict(data)
        row['JSON_File'] = basename
        all_rows.append(row)
        all_keys.update(row.keys())

    if not all_rows:
        return False

    all_rows.sort(
        key=lambda r: (str(r.get('Timestamp') or ''), str(r.get('JSON_File') or ''))
    )

    keys = all_keys
    cols = _survey_column_order(keys)
    cols = [c for c in cols if c != 'JSON_File']
    cols.append('JSON_File')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Surveys'
    for col_idx, name in enumerate(cols, start=1):
        ws.cell(1, col_idx, value=name)
    for ri, row_data in enumerate(all_rows, start=2):
        for col_idx, name in enumerate(cols, start=1):
            val = row_data.get(name, '')
            if val is None:
                val = ''
            elif isinstance(val, (dict, list)):
                val = json.dumps(val, ensure_ascii=False)
            ws.cell(ri, col_idx, value=val)

    _style_survey_worksheet(ws, cols)
    wb.save(SURVEY_EXCEL_FILE)
    print(f'[Survey] Da tao lai Excel tu {len(all_rows)} file JSON -> {SURVEY_EXCEL_FILE}')
    return True


def _append_survey_row_excel(payload, json_filename):
    """Thêm một dòng vào survey_responses.xlsx cùng thư mục với file JSON."""
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        print('[Survey] openpyxl chua cai — bo qua ghi Excel (pip install openpyxl)')
        return

    # Excel rỗng / không đọc được → xóa và build lại từ toàn bộ survey_*.json (JSON đã lưu trước đó)
    if not _survey_excel_file_ok(SURVEY_EXCEL_FILE):
        try:
            if os.path.isfile(SURVEY_EXCEL_FILE):
                os.remove(SURVEY_EXCEL_FILE)
        except OSError:
            pass
        if rebuild_survey_excel_from_json_files():
            return

    row_data = dict(payload)
    row_data['JSON_File'] = json_filename
    keys = set(row_data.keys())

    if os.path.isfile(SURVEY_EXCEL_FILE):
        wb = load_workbook(SURVEY_EXCEL_FILE)
        ws = wb.active
        existing_headers = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if v is not None and str(v).strip() != '':
                existing_headers.append(str(v).strip())
        if not existing_headers:
            cols = _survey_column_order(keys)
        else:
            cols = list(existing_headers)
            for k in _survey_column_order(keys):
                if k not in cols:
                    cols.append(k)
            for k in sorted(keys):
                if k not in cols:
                    cols.append(k)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Surveys'
        existing_headers = []
        cols = _survey_column_order(keys)

    cols = [c for c in cols if c != 'JSON_File']
    if 'JSON_File' in row_data:
        cols.append('JSON_File')

    # Nếu thêm cột mới: map lại các dòng dữ liệu cũ theo header cũ → cột mới
    if existing_headers and existing_headers != cols:
        for r in range(2, ws.max_row + 1):
            prev = {}
            for ci, h in enumerate(existing_headers, start=1):
                if ci <= ws.max_column:
                    prev[h] = ws.cell(row=r, column=ci).value
            for ci, h in enumerate(cols, start=1):
                val = prev.get(h, '')
                if val is None:
                    val = ''
                ws.cell(row=r, column=ci, value=val)

    for col_idx, name in enumerate(cols, start=1):
        ws.cell(row=1, column=col_idx, value=name)
    next_row = ws.max_row + 1
    for col_idx, name in enumerate(cols, start=1):
        val = row_data.get(name, '')
        if val is None:
            val = ''
        elif isinstance(val, (dict, list)):
            val = json.dumps(val, ensure_ascii=False)
        ws.cell(row=next_row, column=col_idx, value=val)

    _style_survey_worksheet(ws, cols)
    wb.save(SURVEY_EXCEL_FILE)
    print(f'[Survey] Excel updated: {SURVEY_EXCEL_FILE} (row {next_row})')


def _hr_save_analysis_snapshot(name, cid, job_title, entry_payload):
    """Lưu output/CV/Analysis_<Tên ứng viên>.json — đồng bộ nội dung với pool HR."""
    try:
        os.makedirs(ANALYSIS_DIR, exist_ok=True)
        base = _safe_analysis_filename(name)
        path = os.path.join(ANALYSIS_DIR, f'Analysis_{base}.json')
        if os.path.isfile(path):
            short = re.sub(r'[^0-9a-fA-F-]', '', (cid or ''))[:10] or 'dup'
            path = os.path.join(ANALYSIS_DIR, f'Analysis_{base}_{short}.json')
        filename = os.path.basename(path)
        out = {
            'candidateId': cid,
            'candidateName': name,
            'jobTitle': job_title or '',
            'savedAt': datetime.now().isoformat(),
            'filename': filename,
            'score': entry_payload.get('score'),
            'recommendation': entry_payload.get('recommendation'),
            'aiStrengths': entry_payload.get('aiStrengths') or [],
            'aiDevelopment': entry_payload.get('aiDevelopment') or [],
            'aiAnalysisSummary': entry_payload.get('aiAnalysisSummary') or '',
            'aiAnalysisImprovement': entry_payload.get('aiAnalysisImprovement') or '',
        }
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(out, f, ensure_ascii=False, indent=2)
        rel = os.path.relpath(path, BASE_DIR)
        print(f'[Server] Saved {rel}')
        return rel.replace('\\', '/')
    except Exception as e:
        print(f'[Server] Analysis snapshot save failed: {e}')
        return None


def _hr_load_pool():
    if not os.path.isfile(HR_POOL_FILE):
        return {'jobTitle': '', 'candidates': []}
    try:
        with open(HR_POOL_FILE, 'r', encoding='utf-8') as f:
            pool = json.load(f)
    except Exception:
        return {'jobTitle': '', 'candidates': []}
    if 'candidates' not in pool:
        pool['candidates'] = []
    pool.pop('jdId', None)
    return pool


def _hr_save_pool(pool):
    pool = dict(pool)
    pool.pop('jdId', None)
    if 'candidates' not in pool:
        pool['candidates'] = []
    with open(HR_POOL_FILE, 'w', encoding='utf-8') as f:
        json.dump(pool, f, ensure_ascii=False, indent=2)

# Anthropic — chỉ từ env ANTHROPIC_API_KEY hoặc file anthropic_key.txt (không lưu key trong repo)
_DEFAULT_ANTHROPIC_KEY = ''

try:
    import shap
    _shap_available = True
except ImportError:
    shap = None
    _shap_available = False
    print('[Server] WARNING: shap not installed — pip install shap')

model = None
feature_names = []
tfidf = None
shap_explainer = None

try:
    model = joblib.load(os.path.join(MODEL_DIR, 'xgb_model.pkl'))
    feature_names = joblib.load(os.path.join(PREPROCESS_DIR, 'feature_names_full.pkl'))
    tfidf = joblib.load(os.path.join(PREPROCESS_DIR, 'tfidf_vectorizer.pkl'))
    print(f'[Server] XGBoost model loaded! Features: {len(feature_names)}')
    print(f'[Server] TF-IDF vectorizer loaded!')
    if _shap_available and model is not None:
        shap_explainer = shap.TreeExplainer(model)
        print('[Server] SHAP TreeExplainer ready!')
except Exception as e:
    print(f'[Server] WARNING: Could not load model: {e}')

# Feature categories (new layout: cv_*, jd_*, CV_JD_Similarity, Gender_*, Race_*, Job Roles_*, Age_Scaled)
CV_TEXT_FEATURES = [f for f in feature_names if f.startswith('cv_')]
JD_TEXT_FEATURES = [f for f in feature_names if f.startswith('jd_')]
GENDER_FEATURES = [f for f in feature_names if f.startswith('Gender_')]
RACE_FEATURES = [f for f in feature_names if f.startswith('Race_')]
JOB_FEATURES = [f for f in feature_names if f.startswith('Job Roles_')]


# ========================================
# ROUTE: Luu CV JSON (tu server.py cu)
# ========================================
def _safe_output_filename(requested):
    """Chỉ lấy basename, chặn path traversal."""
    if not requested or not isinstance(requested, str):
        return None
    base = os.path.basename(requested.strip())
    if not base or '..' in base:
        return None
    if '/' in base or '\\' in base:
        return None
    if not base.lower().endswith('.json'):
        base = base + '.json'
    return base


@app.route('/api/cv/save', methods=['POST'])
def save_cv():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No data'}), 400

        requested = data.get('filename') or (data.get('_meta') or {}).get('jsonFileName')
        filename = _safe_output_filename(requested)

        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            name = (data.get('basicInfo') or {}).get('name', 'unknown')
            name = str(name).replace(' ', '_').replace('/', '_')[:80]
            filename = f'cv_{name}_{timestamp}.json'

        filepath = os.path.join(CV_OUTPUT_DIR, filename)

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        print(f'[Server] Saved CV: {filepath}')
        return jsonify({
            'success': True,
            'filename': filename,
            'path': os.path.abspath(filepath)
        })

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ========================================
# ROUTE: Du doan bang XGBoost
# (tu backend.py - POST /predict)
# ========================================
@app.route('/api/predict', methods=['POST'])
def predict():
    if model is None:
        return jsonify({
            'success': False,
            'message': 'Model chua duoc load. Kiem tra model/xgb_model.pkl va preprocess/feature_names_full.pkl.'
        }), 500

    try:
        data = request.get_json()
        if not data or 'features' not in data:
            return jsonify({'success': False, 'message': 'Thieu truong features'}), 400

        input_data = np.array(data['features'], dtype=float).reshape(1, -1)

        if input_data.shape[1] != len(feature_names):
            return jsonify({
                'success': False,
                'message': f'So features khong khop: gui {input_data.shape[1]}, model can {len(feature_names)}'
            }), 400

        prediction = model.predict(input_data)[0]
        probability = model.predict_proba(input_data)[0].tolist()
        score = round(probability[1] * 100, 1)

        print(f'[Server] Predict: result={int(prediction)}, score={score}%')

        return jsonify({
            'success': True,
            'prediction': int(prediction),   # 1 = Dau, 0 = Rot
            'probability': probability,
            'score': score,                   # % phu hop (0-100)
            'message': 'Phan tich boi XGBoost thanh cong!'
        })

    except Exception as e:
        print(f'[Server] Predict error: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500


# ========================================
# ROUTE: Lay danh sach feature names
# ========================================
@app.route('/api/features', methods=['GET'])
def get_features():
    return jsonify({
        'success': True,
        'features': list(feature_names),
        'count': len(feature_names)
    })

@app.route('/api/chat', methods=['POST'])
def chat():
    data = request.get_json() or {}
    message = data.get('message', '')
    context_raw = data.get('context', '')

    ctx_obj = {}
    if context_raw:
        try:
            ctx_obj = json.loads(context_raw) if isinstance(context_raw, str) else context_raw
        except (json.JSONDecodeError, TypeError, ValueError):
            ctx_obj = {}
    if not isinstance(ctx_obj, dict):
        ctx_obj = {}

    chat_mode = str((ctx_obj.get('chatMode') or ctx_obj.get('chat_mode') or '')).upper()
    ctx_json = json.dumps(ctx_obj, ensure_ascii=False) if ctx_obj else '{}'

    api_key = get_anthropic_key()

    if chat_mode == 'APP_HELP':
        prompt = f"""Bạn là trợ lý hướng dẫn trong hệ thống AI CV Screening / tuyển dụng (tiếng Việt).

NHIỆM VỤ: Trả lời về cách dùng hệ thống, quy trình tuyển dụng dưới đây, và các bước trên giao diện.
Nếu context không có dữ liệu phân tích mà câu hỏi cần điểm số chi tiết — hướng dẫn hoàn tất tải CV, kiểm tra thông tin và chạy phân tích.

QUY TRÌNH TUYỂN DỤNG (chuẩn mô tả cho ứng viên):
1) Ứng viên đưa CV lên hệ thống AI CV Screening.
2) Ứng viên kiểm tra lại một lần: thông tin hệ thống trích xuất có khớp với CV đã tải không; chỉnh sửa trên giao diện nếu lệch.
3) AI đánh giá độ phù hợp CV với JD và các yêu cầu khác của tổ chức; KHÔNG tiết lộ chi tiết tiêu chí nội bộ hay cách tính nội bộ cho ứng viên.
4) Hệ thống trả về điểm mạnh, điểm yếu và hướng cần cải thiện. Phân tích được trình bày trực quan, minh bạch; thiết kế hướng tới không phân biệt và không thiên kiến thuật toán (giải thích ngắn nếu được hỏi về công bằng).
5) HR rà soát thêm; phản hồi/kết quả chính thức thường được gửi qua email trong khoảng 2–3 ngày (có thể nêu khoảng thời gian này khi hỏi sau khi có đánh giá tự động).

QUY TẮC:
- Luôn tiếng Việt, lịch sự, xưng hô "bạn". Gọi là «hệ thống» / «hệ thống này», tránh gọi là «demo» trừ khi người dùng tự nói.
- Không markdown (#, **). Hạn chế emoji. 3–8 câu trừ khi cần liệt kê ngắn.
- Không bịa mức lương, tên công ty cụ thể, hoặc tiêu chí nội bộ không có trong dữ liệu.
- Câu hỏi hoàn toàn ngoài phạm vi tuyển dụng/hệ thống: từ chối ngắn.

LUỒNG MÀN HÌNH (khớp giao diện):
- Màn 0: Chọn «Ứng viên» hoặc «HR».
- Màn 1: Tải CV (PDF); hệ thống trích xuất thông tin.
- Màn 2: Xem và chỉnh thông tin trích xuất; gửi để phân tích / matching.
- Màn 3: Chờ phân tích.
- Màn 4: Xem điểm và gợi ý (điểm mạnh/yếu/cải thiện) nếu đã có kết quả.
- Màn 5: Khảo sát trải nghiệm (nếu có).

LUỒNG HR: Nhập hoặc tải JD; xem danh sách ứng viên theo tab trên giao diện. HR có thể xem/xử lý hồ sơ phù hợp quy trình; kết quả cuối cho ứng viên vẫn thông qua bước rà soát và kênh email như trên.

User context (JSON — có currentScreen, hasUploadedCv, hasAnalysisResult, v.v.):
{ctx_json}

Câu hỏi của người dùng:
{message}

Trả lời:"""
        max_tokens = 1200
    else:
        context = context_raw if isinstance(context_raw, str) else ctx_json
        prompt = f"""Bạn là trợ lý tư vấn trong hệ thống AI CV Screening / tuyển dụng (tiếng Việt).

BỐI CẢNH QUY TRÌNH (ôn lại khi hợp lý): ứng viên tải CV → kiểm tra thông tin trích xuất → AI đánh giá phù hợp CV–JD và tiêu chí nội bộ (không tiết lộ chi tiết tiêu chí) → trả điểm mạnh/yếu/cải thiện → HR rà soát; phản hồi chính thức thường qua email trong khoảng 2–3 ngày. Phân tích nhằm trực quan, minh bạch; không khuyến khích phân biệt hay thiên kiến thuật toán.

QUY TẮC:
- Luôn trả lời bằng tiếng Việt, giọng lịch sự, rõ ràng; xưng hô với người dùng là "bạn".
- Không dùng markdown (#, **). Hạn chế emoji.
- Câu ngắn, dễ đọc. Không giọng đùa, không xưng "em/mình" với ứng viên.
- Chỉ dùng dữ liệu trong User context (điểm, strengths, SHAP, v.v.) khi tư vấn về CV/điểm. Không bịa lương hay tiêu chí nội bộ cụ thể.
- Không liệt kê chi tiết "yêu cầu khác của công ty" nếu không có trong context; có thể nói ngắn là có thêm xét theo quy trình nội bộ.
- Nếu trong context có currentScreen / hasAnalysisResult / hasUploadedCv, có thể nhắc ngắn bước trên giao diện.
- Nếu thiếu dữ liệu điểm/CV trong context mà câu hỏi cần dữ liệu đó, nói rõ và không suy diễn.

User context (JSON: điểm, gợi ý, và có thể có currentScreen, hasUploadedCv, hasAnalysisResult):
{context}

Câu hỏi của người dùng:
{message}

Trả lời:"""
        max_tokens = 5000

    try:
        body = json.dumps({
            "model": "claude-haiku-4-5",
            "max_tokens": max_tokens,
            "messages": [
                {"role": "user", "content": prompt}
            ]
        }).encode("utf-8")

        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=body,
            headers={
                "x-api-key": api_key,
                "content-type": "application/json",
                "anthropic-version": "2023-06-01",
            },
            method="POST",
        )

        with urllib.request.urlopen(req) as res:
            result = json.loads(res.read().decode())

        reply = result["content"][0]["text"]
        reply = clean_text(reply)

        return jsonify({
            "success": True,
            "reply": reply
        })

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
# ========================================
# ROUTE: CV Upload (placeholder cho production)
# ========================================
def clean_text(text):
    if not text:
        return ""

    # Không gộp khoảng trắng giữa hai chữ cái — regex đó làm dính cả từ tiếng Việt
    # ("mình chưa" -> "mìnhchưa"). Chỉ gom khoảng trắng thừa.

    # fix multiple spaces
    text = re.sub(r'\s{2,}', ' ', text)

    # remove markdown
    text = re.sub(r'[#*_`]+', '', text)

    return text.strip()
def get_anthropic_key():
    k = os.environ.get('ANTHROPIC_API_KEY', '').strip()
    if k:
        return k
    keyfile = os.path.join(BASE_DIR, 'anthropic_key.txt')
    if os.path.isfile(keyfile):
        with open(keyfile, 'r', encoding='utf-8') as f:
            k = f.read().strip()
            if k:
                return k
    return _DEFAULT_ANTHROPIC_KEY.strip()


def extract_pdf_text_bytes(data: bytes) -> str:
    text = ''
    if pdfplumber:
        try:
            with pdfplumber.open(io.BytesIO(data)) as pdf:
                parts = [(p.extract_text() or '') for p in pdf.pages]
            text = '\n'.join(parts).strip()
            print(f'[Server] pdfplumber extracted {len(text)} chars from {len(parts)} pages')
        except Exception as e:
            print(f'[Server] pdfplumber error: {e}')
            text = ''
    if not text:
        try:
            import PyPDF2
            reader = PyPDF2.PdfReader(io.BytesIO(data))
            text = '\n'.join((page.extract_text() or '') for page in reader.pages).strip()
            print(f'[Server] PyPDF2 fallback extracted {len(text)} chars')
        except Exception as e:
            print(f'[Server] PyPDF2 fallback error: {e}')
    if not text:
        print(f'[Server] WARNING: No text extracted from PDF ({len(data)} bytes)')
    return text


def _clean_skill_list(val):
    """Không giới hạn số mục — giữ toàn bộ skill hợp lệ."""
    if not val:
        return []
    if isinstance(val, str):
        return [x.strip() for x in re.split(r'[\n,;]', val) if x.strip()]
    if isinstance(val, list):
        out = []
        for x in val:
            if isinstance(x, str) and x.strip():
                out.append(x.strip())
            elif isinstance(x, dict) and (x.get('name') or x.get('skill')):
                out.append(str(x.get('name') or x.get('skill')).strip())
        return out
    return []


def _normalize_skill_categories(r: dict) -> dict:
    sc = r.get('skillCategories') or r.get('skill_categories')
    if isinstance(sc, dict):
        return {
            'languages': _clean_skill_list(sc.get('languages')),
            'tools': _clean_skill_list(sc.get('tools')),
            'hardSkills': _clean_skill_list(sc.get('hardSkills')),
            'softSkills': _clean_skill_list(sc.get('softSkills')),
        }
    skills_raw = r.get('skills') or []
    flat = []
    for s in skills_raw:
        if isinstance(s, str) and s.strip():
            flat.append(s.strip())
        elif isinstance(s, dict):
            nm = str(s.get('name') or s.get('skill') or '').strip()
            if nm:
                flat.append(nm)
    return {
        'languages': [],
        'tools': [],
        'hardSkills': flat,
        'softSkills': [],
    }


def _normalize_upload_cv(raw: dict) -> dict:
    r = raw or {}
    bi = r.get('basicInfo') or r.get('basic_info') or {}
    basic_info = {
        'name': str(bi.get('name', '')).strip(),
        'email': str(bi.get('email', '')).strip(),
        'phone': str(bi.get('phone', '')).strip(),
    }
    skill_categories = _normalize_skill_categories(r)

    exp = r.get('experience')
    if isinstance(exp, list):
        blocks = []
        for block in exp:
            if isinstance(block, str):
                blocks.append(block.strip())
            else:
                header = ' | '.join(
                    x for x in [block.get('duration'), block.get('title'), block.get('company')] if x
                )
                lines = [header]
                if block.get('location'):
                    lines.append(str(block['location']))
                if block.get('description'):
                    lines.append(str(block['description']))
                tech = block.get('technologies')
                if tech:
                    lines.append('Technologies: ' + ', '.join(tech))
                blocks.append('\n'.join(lines))
        exp_str = '\n\n'.join(blocks)
    else:
        exp_str = str(exp or '').strip()

    edu = r.get('education')
    if isinstance(edu, list):
        parts = []
        for ed in edu:
            if isinstance(ed, str):
                parts.append(ed.strip())
            elif isinstance(ed, dict):
                inst = ed.get('institution') or ed.get('school') or ''
                yr = ed.get('graduationYear') or ed.get('year') or ''
                rest = ', '.join(x for x in [ed.get('degree'), ed.get('field')] if x)
                parts.append(' — '.join(x for x in [inst, rest, str(yr)] if x))
        edu_str = '; '.join(parts)
    else:
        edu_str = str(edu or '').strip()

    return {
        'basicInfo': basic_info,
        'skillCategories': skill_categories,
        'experience': exp_str,
        'education': edu_str,
    }


def _quick_doc_check(raw_text: str, expected: str, api_key: str, pdf_bytes: bytes = None) -> bool:
    """Hỏi Claude 1 câu đơn giản: đây có phải là CV/JD không? Trả True nếu đúng.
    Nếu raw_text rỗng nhưng có pdf_bytes, gửi PDF trực tiếp cho Claude đọc."""
    if expected == 'cv':
        question = 'Is this document a CV or Resume of a person? Answer ONLY "yes" or "no".'
    else:
        question = 'Is this document a Job Description or job posting? Answer ONLY "yes" or "no".'

    if raw_text and raw_text.strip():
        snippet = raw_text[:3000]
        user_content = f'{question}\n\n{snippet}'
    elif pdf_bytes:
        user_content = [
            {
                'type': 'document',
                'source': {
                    'type': 'base64',
                    'media_type': 'application/pdf',
                    'data': base64.standard_b64encode(pdf_bytes).decode('ascii'),
                },
            },
            {'type': 'text', 'text': question},
        ]
    else:
        return True

    body = json.dumps({
        'model': 'claude-haiku-4-5',
        'max_tokens': 10,
        'messages': [{'role': 'user', 'content': user_content}],
    }).encode('utf-8')
    req = urllib.request.Request(
        'https://api.anthropic.com/v1/messages',
        data=body,
        headers={
            'Content-Type': 'application/json',
            'x-api-key': api_key,
            'anthropic-version': '2023-06-01',
            'anthropic-beta': 'pdfs-2024-09-25',
        },
        method='POST',
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            out = json.loads(resp.read().decode('utf-8'))
        answer = out['content'][0]['text'].strip().lower()
        print(f'[Server] Quick doc check ({expected}): "{answer}"')
        return answer.startswith('yes')
    except Exception as e:
        print(f'[Server] Quick doc check error: {e} — allowing by default')
        return True


def call_claude_extract(raw_text: str, api_key: str, pdf_bytes: bytes = None) -> dict:
    """Trích CV qua Claude. Nếu raw_text rỗng nhưng có pdf_bytes, gửi PDF trực tiếp."""
    instructions = (
        'You are a CV data extractor. Read the CV and return ONLY valid JSON '
        '(no markdown fences, no commentary).\n\nRules:\n'
        '- IMPORTANT: Keep the ORIGINAL language of the CV. If the CV is in Vietnamese, output Vietnamese. If in English, output English. Do NOT translate.\n'
        '- Plain text only; no odd symbols; use simple line breaks.\n'
        '- JSON has EXACTLY 4 keys (no "skills" key — only "skillCategories"):\n'
        '1) "basicInfo": { "name", "email", "phone" } — use "" if missing.\n'
        '3) "skillCategories": object with FOUR arrays of strings. '
        'Include EVERY skill mentioned anywhere in the CV. Do not cap or omit. Use [] only if nothing.\n'
        '   - "languages": spoken/written languages with proficiency if stated (e.g. "English — Professional").\n'
        '   - "tools": software, IDEs, Git, cloud platforms, BI tools, Office apps, design tools, OS.\n'
        '   - "hardSkills": programming languages, frameworks, testing methods, domain skills, methodologies.\n'
        '   - "softSkills": teamwork, leadership, communication, problem-solving, etc.\n'
        '4) "experience": ONE string — all jobs concatenated; '
        'format: "Dates\\nCompany — Title\\nDescription". Blank line between companies.\n'
        '5) "education": ONE short string — ONLY school names, degree/major, and graduation years. '
        'Separate multiple entries with "; ". '
        'Do NOT list individual course names or subjects.\n\n'
    )

    use_pdf = (not raw_text or not raw_text.strip()) and pdf_bytes
    headers = {
        'Content-Type': 'application/json',
        'x-api-key': api_key,
        'anthropic-version': '2023-06-01',
    }

    if use_pdf:
        print(f'[Server] call_claude_extract: using PDF native mode ({len(pdf_bytes)} bytes)')
        headers['anthropic-beta'] = 'pdfs-2024-09-25'
        user_content = [
            {
                'type': 'document',
                'source': {
                    'type': 'base64',
                    'media_type': 'application/pdf',
                    'data': base64.standard_b64encode(pdf_bytes).decode('ascii'),
                },
            },
            {'type': 'text', 'text': instructions + 'Extract from the PDF above.'},
        ]
    else:
        text_slice = raw_text if len(raw_text) <= 90000 else raw_text[:90000]
        user_content = instructions + 'TEXT:\n' + text_slice

    body = json.dumps({
        'model': 'claude-haiku-4-5',
        'max_tokens': 8192,
        'messages': [{'role': 'user', 'content': user_content}],
    }).encode('utf-8')
    req = urllib.request.Request(
        'https://api.anthropic.com/v1/messages',
        data=body,
        headers=headers,
        method='POST',
    )
    try:
        with urllib.request.urlopen(req, timeout=180) as resp:
            out = json.loads(resp.read().decode('utf-8'))
    except urllib.error.HTTPError as e:
        err = e.read().decode('utf-8', errors='replace')[:500]
        raise RuntimeError(f'Claude API HTTP {e.code}: {err}') from e
    text = out['content'][0]['text'].strip()
    text = re.sub(r'^```json\s*', '', text, flags=re.I)
    text = re.sub(r'^```\s*', '', text)
    text = re.sub(r'\s*```$', '', text)
    return json.loads(text.strip())

@app.route('/api/survey/submit', methods=['POST', 'OPTIONS'])
def submit_survey():
    if request.method == 'OPTIONS':
        response = jsonify({'success': True})
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type')
        response.headers.add('Access-Control-Allow-Methods', 'POST')
        return response
    try:
        data = request.get_json() or {}
        print('[Survey] Received:', data)
        name_raw = (data.get('Candidate_Name') or data.get('candidate_name') or '').strip() or 'Unknown'
        # Cùng stem với CV_<stem>.json; phần trùng file giống Analysis_<base>_<suffix>.json
        stem = _safe_cv_name_stem(name_raw)
        survey_file = os.path.join(SURVEY_OUTPUT_DIR, f'survey_{stem}.json')
        if os.path.isfile(survey_file):
            cid = (data.get('CV_ID') or '').strip()
            short = re.sub(r'[^0-9a-zA-Z_\-]', '', cid)[-10:] if cid else ''
            if short:
                survey_file = os.path.join(SURVEY_OUTPUT_DIR, f'survey_{stem}_{short}.json')
            else:
                survey_file = os.path.join(
                    SURVEY_OUTPUT_DIR,
                    f'survey_{stem}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
                )
        saved_name = os.path.basename(survey_file)
        with open(survey_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f'[Survey] Saved to {survey_file}')
        try:
            _append_survey_row_excel(data, saved_name)
        except Exception as ex:
            print(f'[Survey] Excel append failed: {ex}')
        return jsonify({
            'success': True,
            'message': 'Survey saved',
            'filename': saved_name,
        })
    except Exception as e:
        print(f'[Survey] Error: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500
@app.route('/api/cv/upload', methods=['POST'])
def upload_cv():
    """Nhan PDF -> trich text -> Claude -> luu JSON output/ -> tra ve data cho Form."""
    print(f'[Server] /api/cv/upload called, files: {list(request.files.keys())}')
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'Thieu file (field name: file)'}), 400
    up = request.files['file']
    print(f'[Server] File received: {up.filename}, content_type: {up.content_type}')
    if not up.filename:
        return jsonify({'success': False, 'message': 'Ten file rong'}), 400

    raw_bytes = up.read()
    print(f'[Server] CV: Read {len(raw_bytes)} bytes')
    raw_text = extract_pdf_text_bytes(raw_bytes)
    print(f'[Server] CV: Extracted text = {len(raw_text)} chars, first 100: {repr(raw_text[:100])}')

    text_empty = not raw_text or not raw_text.strip()
    if text_empty and len(raw_bytes) < 500:
        return jsonify({
            'success': False,
            'message': 'File PDF trống hoặc bị hỏng. Vui lòng chọn file khác.',
            'errorType': 'unreadable_pdf'
        }), 400

    if text_empty:
        print('[Server] CV: Text extraction failed — will use Claude PDF native mode')

    api_key = get_anthropic_key()
    if not api_key:
        return jsonify({
            'success': False,
            'message': 'Thieu ANTHROPIC_API_KEY (env) hoac file anthropic_key.txt trong thu muc project',
        }), 503

    pdf_for_claude = raw_bytes if text_empty else None
    if not _quick_doc_check(raw_text, 'cv', api_key, pdf_bytes=pdf_for_claude):
        return jsonify({
            'success': False,
            'message': 'File bạn tải lên không phải là CV. Vui lòng chọn đúng file hồ sơ cá nhân (CV/Resume).',
            'errorType': 'wrong_document'
        }), 400

    try:
        parsed = call_claude_extract(raw_text, api_key, pdf_bytes=pdf_for_claude)
        data = _normalize_upload_cv(parsed)
    except Exception as e:
        print('[Server] /api/cv/upload error:', e)
        return jsonify({'success': False, 'message': str(e)}), 500

    if text_empty:
        bi = data.get('basicInfo', {})
        sc = data.get('skillCategories', {})
        all_skills = ' '.join(
            sc.get('languages', []) + sc.get('tools', []) +
            sc.get('hardSkills', []) + sc.get('softSkills', [])
        )
        raw_text = ' '.join(filter(None, [
            bi.get('name', ''), all_skills,
            data.get('experience', ''), data.get('education', ''),
        ]))
        print(f'[Server] CV: Reconstructed raw_text from Claude extraction ({len(raw_text)} chars)')

    cv_id = f'cv_{int(datetime.now().timestamp())}'
    safe_name = _safe_cv_name_stem(data['basicInfo'].get('name'))
    requested_fn = f'CV_{safe_name}.json'
    filename = _safe_output_filename(requested_fn) or f'cv_{cv_id}.json'
    filepath = os.path.join(CV_OUTPUT_DIR, filename)

    payload = {
        'filename': filename,
        'basicInfo': data['basicInfo'],
        'skillCategories': data.get('skillCategories'),
        'experience': data['experience'],
        'education': data['education'],
        '_meta': {
            'cvId': cv_id,
            'extractedAt': datetime.now().isoformat(),
            'rawTextLength': len(raw_text),
        },
    }
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f'[Server] /api/cv/upload saved: {filepath}')

    return jsonify({
        'success': True,
        'cvId': cv_id,
        'data': data,
        'rawText': raw_text,
        'saved': True,
        'filename': filename,
        'path': os.path.abspath(filepath),
    })


# ========================================
# ROUTE: Analysis Result
# ========================================
@app.route('/api/analysis/<analysis_id>', methods=['GET'])
def get_analysis(analysis_id):
    return jsonify({
        'success': True,
        'analysisId': analysis_id,
        'status': 'completed',
        'message': 'Ket qua da duoc luu trong AppState'
    })


# ========================================
# ROUTE: CV text -> vector 721 chieu (TF-IDF + one-hot + tuoi)
# TfidfVectorizer (joblib) da hoc tokenizer + vocabulary — chi can 1 chuoi text day du.
# ========================================
@app.route('/api/preprocess', methods=['POST'])
def preprocess():
    if tfidf is None:
        return jsonify({'success': False, 'message': 'TF-IDF chua duoc load'}), 500

    try:
        data = request.get_json()
        raw_text = data.get('text', '')
        jd_text = data.get('jdText', '')
        job_role = data.get('jobRole', '')
        gender = data.get('gender', '')
        race = data.get('race', '')
        age = data.get('age', 28)

        input_arr, _meta = _build_features(raw_text, jd_text, job_role, gender, race, age)
        features = input_arr.flatten().tolist()
        nonzero = sum(1 for v in features if v != 0)
        print(f'[Server] Preprocess: {len(features)} features, {nonzero} non-zero')

        return jsonify({
            'success': True,
            'features': features,
            'feature_count': len(features),
            'nonzero_count': nonzero
        })

    except Exception as e:
        print(f'[Server] Preprocess error: {e}')
        return jsonify({'success': False, 'message': str(e)}), 500


def _normalize_text(s):
    """Lowercase, bỏ dấu câu, chuẩn hóa khoảng trắng."""
    s = re.sub(r'[^\w\s]', ' ', str(s).lower())
    return ' '.join(s.split())


_VI_EN_MAP = {
    'phân tích': 'analysis analytical analyze',
    'dữ liệu': 'data',
    'quản lý': 'management manage',
    'kỹ năng': 'skills skill',
    'giao tiếp': 'communication',
    'lãnh đạo': 'leadership',
    'làm việc nhóm': 'teamwork team',
    'giải quyết vấn đề': 'problem solving',
    'lập kế hoạch': 'planning',
    'nghiên cứu': 'research',
    'thiết kế': 'design',
    'phát triển': 'development develop',
    'kiểm thử': 'testing test',
    'tối ưu': 'optimization optimize',
    'báo cáo': 'reporting report',
    'tài chính': 'financial finance',
    'khách hàng': 'customer client',
    'hệ thống': 'systems system',
    'cơ sở dữ liệu': 'database',
    'bảo mật': 'security',
    'triển khai': 'deploy implementation',
    'quy trình': 'process',
    'yêu cầu': 'requirements',
    'đào tạo': 'training',
    'viết lách': 'writing',
    'sáng tạo': 'creative',
    'truyền thông': 'media communication',
    'quảng cáo': 'advertising',
    'thương hiệu': 'branding brand',
    'nội dung': 'content',
    'sự kiện': 'event',
    'chiến lược': 'strategy strategic',
    'marketing': 'marketing',
    'đại học': 'bachelor university',
}


def _expand_vi_to_en(text_lower: str) -> str:
    """Mở rộng text bằng cách thêm từ tiếng Anh tương ứng cho các cụm tiếng Việt."""
    expanded = text_lower
    for vi, en in _VI_EN_MAP.items():
        if vi in text_lower:
            expanded += ' ' + en
    return expanded


_STOP_WORDS = frozenset({
    'the', 'and', 'for', 'with', 'that', 'this', 'from', 'are', 'was',
    'were', 'been', 'being', 'have', 'has', 'had', 'will', 'would',
    'could', 'should', 'may', 'might', 'shall', 'can', 'not', 'but',
    'its', 'our', 'your', 'their', 'his', 'her', 'all', 'any', 'each',
    'such', 'into', 'over', 'also', 'than', 'then', 'when', 'where',
    'which', 'who', 'how', 'what', 'both', 'use', 'used', 'using',
    'including', 'includes', 'include', 'within', 'between', 'during',
    'about', 'under', 'through', 'before', 'after', 'other', 'some',
})


def _extract_tech_terms(text: str) -> set:
    """Trích thuật ngữ kỹ thuật: viết hoa, viết tắt, tên tool.
    Lọc bỏ stop words để tránh false positive matches."""
    terms = set()
    for word in re.findall(r'[A-Za-z][A-Za-z0-9+#.]{1,}', text):
        lower = word.lower()
        if len(lower) >= 3 and lower not in _STOP_WORDS:
            terms.add(lower)
    return terms


def _build_prefix_set(words, min_len=4):
    """Build a set of word prefixes for fuzzy matching (handles plurals, -ing, -tion etc.)."""
    prefixes = set()
    for w in words:
        if len(w) >= min_len:
            prefixes.add(w[:min_len])
            if len(w) >= 5:
                prefixes.add(w[:5])
    return prefixes


def _fuzzy_word_hit(word, exact_set, prefix_set, min_prefix=4):
    """Check if a word matches via exact match or shared prefix."""
    if word in exact_set:
        return True
    if len(word) >= min_prefix and word[:min_prefix] in prefix_set:
        return True
    return False


def _compute_skill_match(cv_items: list, jd_items: list) -> dict:
    """So sánh trực tiếp CV skills vs JD requirements.
    Uses prefix matching to handle morphological variations
    (plurals, -ing, -tion, -ment etc.) and confidence-weighted scoring."""
    if not jd_items:
        return {'match_ratio': 0.5, 'matched': 0, 'total_jd': 0,
                'matched_details': [], 'unmatched_details': []}

    cv_blob_raw = ' '.join(str(s) for s in cv_items)
    cv_blob = _normalize_text(cv_blob_raw)
    cv_expanded = _expand_vi_to_en(cv_blob)
    cv_words = set(w for w in cv_expanded.split() if len(w) >= 3 and w not in _STOP_WORDS)
    cv_tech = _extract_tech_terms(cv_blob_raw)
    cv_prefixes = _build_prefix_set(cv_words)
    cv_tech_prefixes = _build_prefix_set(cv_tech)

    matched_details = []
    unmatched_details = []

    for jd_item in jd_items:
        jd_str = str(jd_item).strip()
        if len(jd_str) < 5:
            continue

        jd_norm = _normalize_text(jd_str)
        jd_expanded = _expand_vi_to_en(jd_norm)
        jd_tech = _extract_tech_terms(jd_str)

        jd_en_words = set(w for w in jd_expanded.split()
                          if len(w) >= 3 and w.isascii() and w not in _STOP_WORDS)
        if not jd_en_words and not jd_tech:
            continue

        en_hits = sum(1 for kw in jd_en_words
                      if _fuzzy_word_hit(kw, cv_words, cv_prefixes)) if jd_en_words else 0
        en_ratio = en_hits / len(jd_en_words) if jd_en_words else 0

        tech_hits = sum(1 for t in jd_tech
                        if _fuzzy_word_hit(t, cv_tech, cv_tech_prefixes)) if jd_tech else 0
        tech_ratio = tech_hits / len(jd_tech) if jd_tech else 0

        confidence = max(en_ratio, tech_ratio)

        if confidence >= 0.20:
            matched_details.append({'item': jd_str, 'confidence': round(confidence, 2)})
        else:
            unmatched_details.append({'item': jd_str, 'confidence': round(confidence, 2)})

    total = len(matched_details) + len(unmatched_details)
    if total == 0:
        return {'match_ratio': 0.5, 'matched': 0, 'total_jd': 0,
                'matched_details': [], 'unmatched_details': []}

    # Confidence-weighted ratio: high-confidence matches count more.
    # Use min denominator of 25 so small JDs (few items) don't get
    # inflated scores from a handful of generic skill matches.
    weighted_sum = sum(m['confidence'] for m in matched_details)
    effective_total = max(total, 25)
    match_ratio = min(1.0, (weighted_sum / effective_total) * 2.5)

    matched_details.sort(key=lambda x: x['confidence'], reverse=True)
    unmatched_details.sort(key=lambda x: x['confidence'])

    print(f'[Server] Skill match: {len(matched_details)}/{total} items, '
          f'weighted_ratio={match_ratio:.3f}')
    for m in matched_details[:3]:
        print(f'  ✓ {m["item"][:60]}  conf={m["confidence"]}')
    for u in unmatched_details[:3]:
        print(f'  ✗ {u["item"][:60]}  conf={u["confidence"]}')

    return {
        'match_ratio': round(match_ratio, 3),
        'matched': len(matched_details),
        'total_jd': total,
        'matched_details': matched_details[:10],
        'unmatched_details': unmatched_details[:10],
    }


def _translate_skill_items(skill_match_result: dict, api_key: str) -> dict:
    """Dịch matched_details & unmatched_details sang tiếng Việt qua Claude Haiku.
    Thêm trường 'item_vi' vào mỗi entry. Nếu lỗi thì giữ nguyên."""
    import re as _re
    matched = skill_match_result.get('matched_details', [])
    unmatched = skill_match_result.get('unmatched_details', [])
    all_items = []
    for m in matched:
        all_items.append(m['item'])
    for u in unmatched:
        all_items.append(u['item'])

    if not all_items or not api_key:
        return skill_match_result

    vi_chars = set('àáảãạăắằẳẵặâấầẩẫậèéẻẽẹêếềểễệìíỉĩịòóỏõọôốồổỗộơớờởỡợùúủũụưứừửữựỳýỷỹỵđ')
    def _is_vietnamese(t):
        return any(c in vi_chars for c in t.lower())

    needs_translate = [(i, s) for i, s in enumerate(all_items) if not _is_vietnamese(s)]
    if not needs_translate:
        for m in matched:
            m['item_vi'] = m['item']
        for u in unmatched:
            u['item_vi'] = u['item']
        return skill_match_result

    numbered = '\n'.join(f'{i+1}. {s}' for i, (_, s) in enumerate(needs_translate))
    prompt = (
        'Dịch danh sách bên dưới sang tiếng Việt.\n'
        'QUY TẮC BẮT BUỘC:\n'
        '- Trả về ĐÚNG định dạng: mỗi dòng là "số. bản dịch tiếng Việt"\n'
        '- KHÔNG thêm giải thích, ghi chú, markdown, dấu ** hay bất kỳ ký tự thừa nào\n'
        '- Dịch tự nhiên, ngắn gọn, giữ nghĩa chuyên ngành\n'
        '- Ví dụ đầu vào: "1. Risk management" → đầu ra: "1. Quản lý rủi ro"\n\n'
        f'{numbered}'
    )

    try:
        payload = json.dumps({
            'model': 'claude-haiku-4-5',
            'max_tokens': 2048,
            'messages': [{'role': 'user', 'content': prompt}],
        }).encode('utf-8')

        req = urllib.request.Request(
            'https://api.anthropic.com/v1/messages',
            data=payload,
            headers={
                'x-api-key': api_key,
                'content-type': 'application/json',
                'anthropic-version': '2023-06-01',
            },
            method='POST',
        )

        with urllib.request.urlopen(req, timeout=15) as resp:
            body = json.loads(resp.read().decode('utf-8'))

        reply = ''
        for block in body.get('content', []):
            if block.get('type') == 'text':
                reply += block['text']

        translations = {}
        for line in reply.strip().split('\n'):
            line = line.strip()
            line = _re.sub(r'\*\*', '', line)
            line = _re.sub(r'^#+\s*', '', line)
            line = _re.sub(r'^[-•]\s*', '', line)
            line = line.strip()
            m = _re.match(r'^(\d+)[\.\)]\s*(.+)$', line)
            if m:
                vi_text = m.group(2).strip()
                vi_text = _re.split(r'\s*[=\(]', vi_text)[0].strip()
                if vi_text:
                    idx = int(m.group(1)) - 1
                    translations[idx] = vi_text

        vi_map = {}
        for seq_idx, (orig_idx, orig_str) in enumerate(needs_translate):
            vi_map[orig_idx] = translations.get(seq_idx, orig_str)

        for i, item_str in enumerate(all_items):
            vi_text = vi_map.get(i, item_str)
            if i < len(matched):
                matched[i]['item_vi'] = vi_text
            else:
                unmatched[i - len(matched)]['item_vi'] = vi_text

        print(f'[Server] Translated {len(translations)}/{len(needs_translate)} skill items to Vietnamese')

    except Exception as e:
        print(f'[Server] translate exception: {e}')
        for m in matched:
            if 'item_vi' not in m:
                m['item_vi'] = m['item']
        for u in unmatched:
            if 'item_vi' not in u:
                u['item_vi'] = u['item']

    return skill_match_result


def _build_features(cv_text, jd_text='', job_role='', gender='', race='', age=28):
    """CV TF-IDF + JD TF-IDF + cosine similarity + one-hot + age -> 1058-dim vector.
    Gender, Race, Age are zeroed out to prevent bias.
    Returns (feature_array, metadata_dict) where metadata contains TF-IDF coverage info."""
    from sklearn.metrics.pairwise import cosine_similarity as _cos_sim
    X_cv = tfidf.transform([cv_text.lower()]).toarray()[0]
    X_jd = tfidf.transform([jd_text.lower()]).toarray()[0] if jd_text else np.zeros_like(X_cv)
    cv_sparse = X_cv.reshape(1, -1)
    jd_sparse = X_jd.reshape(1, -1)
    cos_sim = float(_cos_sim(cv_sparse, jd_sparse)[0, 0]) if jd_text else 0.0
    X_gender = [0] * len(GENDER_FEATURES)
    X_race = [0] * len(RACE_FEATURES)
    X_job = [1 if f == f'Job Roles_{job_role}' else 0 for f in JOB_FEATURES]
    age_scaled = 0.0
    features = list(X_cv) + list(X_jd) + [cos_sim] + X_gender + X_race + X_job + [age_scaled]

    cv_nonzero = int(np.count_nonzero(X_cv))
    jd_nonzero = int(np.count_nonzero(X_jd))
    meta = {
        'cv_tfidf_nonzero': cv_nonzero,
        'jd_tfidf_nonzero': jd_nonzero,
        'cosine_sim': round(cos_sim, 4),
    }
    return np.array(features, dtype=np.float64).reshape(1, -1), meta


def _get_shap_explanation(input_arr):
    """Run SHAP on input_arr (1, 721) and return rich explanation dict.
    Separates 'present' features (input > 0) from 'absent' features (input = 0)
    so the UI can show what IS in the CV vs what's missing."""
    if shap_explainer is None:
        return None

    sv = shap_explainer(input_arr)
    sv.feature_names = list(feature_names)
    vals = sv[0]
    input_flat = input_arr.flatten()

    present_pos = []
    present_neg = []
    absent_helpful = []

    for i in range(len(feature_names)):
        v = float(vals.values[i])
        if v == 0:
            continue
        fname = feature_names[i]
        if fname.startswith('Gender_') or fname.startswith('Race_') or fname == 'Age_Scaled':
            continue
        cat = 'text'
        if fname.startswith('Job Roles_'):
            cat = 'job_role'

        entry = {
            'feature': fname,
            'impact': round(v, 6),
            'category': cat,
            'present': bool(input_flat[i] != 0),
        }

        if input_flat[i] != 0:
            if v > 0:
                present_pos.append(entry)
            else:
                present_neg.append(entry)
        else:
            if v > 0.02:
                absent_helpful.append(entry)

    present_pos.sort(key=lambda x: x['impact'], reverse=True)
    present_neg.sort(key=lambda x: x['impact'])
    absent_helpful.sort(key=lambda x: x['impact'], reverse=True)

    all_impacts = [float(vals.values[i]) for i in range(len(feature_names)) if vals.values[i] != 0]
    total_pos = round(sum(v for v in all_impacts if v > 0), 4)
    total_neg = round(sum(v for v in all_impacts if v < 0), 4)

    return {
        'base_score': round(float(vals.base_values), 4),
        'top_positives': present_pos[:10],
        'top_negatives': present_neg[:10],
        'absent_helpful': absent_helpful[:5],
        'total_positive_impact': total_pos,
        'total_negative_impact': total_neg,
        'nonzero_features': len(all_impacts),
        'present_features_count': sum(1 for x in input_flat if x != 0),
    }


# ========================================
# ROUTE: Combined score — preprocess + predict + SHAP in one call
# ========================================
@app.route('/api/score', methods=['POST'])
def score_cv():
    """Nhan text CV -> TF-IDF -> XGBoost predict -> SHAP explanation."""
    if model is None or tfidf is None:
        return jsonify({
            'success': False,
            'message': 'Model hoac TF-IDF chua duoc load'
        }), 500

    try:
        data = request.get_json()
        raw_text = data.get('text', '')
        jd_text = data.get('jdText', '')
        cv_skills = data.get('cvSkills', [])
        jd_skills = data.get('jdSkills', [])
        job_role = data.get('jobRole', '')
        gender = data.get('gender', '')
        race = data.get('race', '')
        age = data.get('age', 28)

        if not raw_text or len(raw_text.strip()) < 10:
            return jsonify({'success': False, 'message': 'Text CV qua ngan'}), 400

        has_jd = bool(jd_text and jd_text.strip()) or bool(jd_skills)
        print(f'[Server] /api/score: CV text={len(raw_text)} chars, JD text={len(jd_text)} chars, '
              f'cvSkills={len(cv_skills)}, jdSkills={len(jd_skills)}')

        # ── XGBoost + SHAP ──
        input_arr, feat_meta = _build_features(raw_text, jd_text, job_role, gender, race, age)
        prediction = int(model.predict(input_arr)[0])
        proba = model.predict_proba(input_arr)[0].tolist()
        xgb_score = float(proba[1])
        explanation = _get_shap_explanation(input_arr)

        # ── Direct skill match (language-agnostic) ──
        skill_match_result = _compute_skill_match(cv_skills, jd_skills)
        skill_ratio = skill_match_result['match_ratio']

        # ── Dịch skill items sang tiếng Việt cho feedback ──
        _translate_skill_items(skill_match_result, get_anthropic_key())

        if has_jd:
            # Adaptive weighting: XGBoost gets more weight when TF-IDF has
            # enough English features to work with; for Vietnamese-heavy input
            # (few TF-IDF hits), skill match dominates.
            total_tfidf_nz = feat_meta['cv_tfidf_nonzero'] + feat_meta['jd_tfidf_nonzero']
            coverage = min(1.0, total_tfidf_nz / 50)
            xgb_weight = 0.7 * coverage
            skill_weight = 1.0 - xgb_weight

            blended = xgb_score * xgb_weight + skill_ratio * skill_weight
            score_pct = round(blended * 100, 1)
            final_pred = 1 if score_pct >= 80 else 0
            print(f'[Server] /api/score: xgb={round(xgb_score*100,1)}%, '
                  f'skill_match={round(skill_ratio*100,1)}% ({skill_match_result["matched"]}/{skill_match_result["total_jd"]}), '
                  f'tfidf_nz={total_tfidf_nz}, coverage={coverage:.2f}, '
                  f'weights=xgb:{xgb_weight:.2f}/skill:{skill_weight:.2f}, '
                  f'blended={score_pct}%')
        else:
            score_pct = round(xgb_score * 100, 1)
            final_pred = prediction
            print(f'[Server] /api/score: no JD — xgb_only={score_pct}%')

        if explanation:
            explanation['skill_match'] = skill_match_result

        return jsonify({
            'success': True,
            'prediction': final_pred,
            'score': score_pct,
            'match_percentage': f'{score_pct}%',
            'probability': proba,
            'xgb_score': round(xgb_score * 100, 1),
            'skill_match': skill_match_result,
            'explanation': explanation,
            'feature_stats': {
                'total': len(feature_names),
                'nonzero_input': int(np.count_nonzero(input_arr)),
                'cv_tfidf_nonzero': feat_meta['cv_tfidf_nonzero'],
                'jd_tfidf_nonzero': feat_meta['jd_tfidf_nonzero'],
                'cosine_sim': feat_meta['cosine_sim'],
            },
        })

    except Exception as e:
        print(f'[Server] /api/score error: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


# ========================================
# ROUTE: JD Upload — extract Job Description from PDF
# ========================================
def call_claude_extract_jd(raw_text: str, api_key: str, pdf_bytes: bytes = None) -> dict:
    """Trích JD qua Claude. Nếu raw_text rỗng nhưng có pdf_bytes, gửi PDF trực tiếp."""
    instructions = (
        'You are a Job Description (JD) data extractor. Read the JD and return ONLY valid JSON '
        '(no markdown fences, no commentary).\n\nRules:\n'
        '- Keep the ORIGINAL language. Do NOT translate.\n'
        '- JSON keys:\n'
        '1) "jobTitle": string — the job title/position/role name. This is REQUIRED.\n'
        '3) "company": string — company name (use "" if missing).\n'
        '4) "location": string — work location (use "" if missing).\n'
        '5) "employmentType": string — full-time, part-time, contract, etc. (use "" if missing).\n'
        '6) "description": string — a brief summary of the role (2-4 sentences).\n'
        '7) "responsibilities": array of strings — main duties/responsibilities.\n'
        '8) "requirements": object with:\n'
        '   - "education": string — required education level.\n'
        '   - "experience": string — required years/type of experience.\n'
        '   - "skills": array of strings — required technical and soft skills.\n'
        '   - "certifications": array of strings — required certifications (use [] if none).\n'
        '9) "preferredQualifications": array of strings — nice-to-have qualifications.\n'
        '10) "benefits": array of strings — listed benefits (use [] if missing).\n'
        '11) "salary": string — salary range if mentioned (use "" if missing).\n\n'
    )

    use_pdf = (not raw_text or not raw_text.strip()) and pdf_bytes
    headers = {
        'Content-Type': 'application/json',
        'x-api-key': api_key,
        'anthropic-version': '2023-06-01',
    }

    if use_pdf:
        print(f'[Server] call_claude_extract_jd: using PDF native mode ({len(pdf_bytes)} bytes)')
        headers['anthropic-beta'] = 'pdfs-2024-09-25'
        user_content = [
            {
                'type': 'document',
                'source': {
                    'type': 'base64',
                    'media_type': 'application/pdf',
                    'data': base64.standard_b64encode(pdf_bytes).decode('ascii'),
                },
            },
            {'type': 'text', 'text': instructions + 'Extract from the PDF above.'},
        ]
    else:
        text_slice = raw_text if len(raw_text) <= 90000 else raw_text[:90000]
        user_content = instructions + 'TEXT:\n' + text_slice

    body = json.dumps({
        'model': 'claude-haiku-4-5',
        'max_tokens': 8192,
        'messages': [{'role': 'user', 'content': user_content}],
    }).encode('utf-8')
    req = urllib.request.Request(
        'https://api.anthropic.com/v1/messages',
        data=body,
        headers=headers,
        method='POST',
    )
    try:
        with urllib.request.urlopen(req, timeout=180) as resp:
            out = json.loads(resp.read().decode('utf-8'))
    except urllib.error.HTTPError as e:
        err = e.read().decode('utf-8', errors='replace')[:500]
        raise RuntimeError(f'Claude API HTTP {e.code}: {err}') from e
    text = out['content'][0]['text'].strip()
    text = re.sub(r'^```json\s*', '', text, flags=re.I)
    text = re.sub(r'^```\s*', '', text)
    text = re.sub(r'\s*```$', '', text)
    return json.loads(text.strip())


@app.route('/api/jd/upload', methods=['POST'])
def upload_jd():
    """Nhan JD PDF -> trich text -> Claude extract -> tra ve JSON."""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'Thieu file (field name: file)'}), 400
    up = request.files['file']
    if not up.filename:
        return jsonify({'success': False, 'message': 'Ten file rong'}), 400

    raw_bytes = up.read()
    raw_text = extract_pdf_text_bytes(raw_bytes)
    print(f'[Server] JD: Extracted text = {len(raw_text)} chars')

    text_empty = not raw_text or not raw_text.strip()
    if text_empty and len(raw_bytes) < 500:
        return jsonify({
            'success': False,
            'message': 'File PDF trống hoặc bị hỏng. Vui lòng chọn file khác.',
            'errorType': 'unreadable_pdf'
        }), 400

    if text_empty:
        print('[Server] JD: Text extraction failed — will use Claude PDF native mode')

    api_key = get_anthropic_key()
    if not api_key:
        return jsonify({
            'success': False,
            'message': 'Thieu ANTHROPIC_API_KEY',
        }), 503

    pdf_for_claude = raw_bytes if text_empty else None
    if not _quick_doc_check(raw_text, 'jd', api_key, pdf_bytes=pdf_for_claude):
        return jsonify({
            'success': False,
            'message': 'File bạn tải lên không phải là nội dung JD. Vui lòng chọn đúng file mô tả công việc.',
            'errorType': 'wrong_document'
        }), 400

    try:
        jd_data = call_claude_extract_jd(raw_text, api_key, pdf_bytes=pdf_for_claude)
    except Exception as e:
        print('[Server] /api/jd/upload error:', e)
        return jsonify({'success': False, 'message': str(e)}), 500

    if text_empty:
        parts = [jd_data.get('jobTitle', ''), jd_data.get('description', '')]
        reqs = jd_data.get('requirements', {})
        if isinstance(reqs, dict):
            parts.extend(reqs.get('skills', []))
        parts.extend(jd_data.get('responsibilities', []))
        raw_text = ' '.join(filter(None, parts))
        print(f'[Server] JD: Reconstructed raw_text from Claude extraction ({len(raw_text)} chars)')

    payload = {
        'filename': 'JD_current.json',
        'jd_data': jd_data,
        'rawText': raw_text,
        '_meta': {
            'extractedAt': datetime.now().isoformat(),
            'rawTextLength': len(raw_text),
        },
    }
    with open(JD_CURRENT_FILE, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    # Lưu bản copy PDF JD vào output/JD/
    try:
        safe_pdf = re.sub(r'[^\w.\-]', '_', (up.filename or 'jd').strip())[:120]
        if not safe_pdf.lower().endswith('.pdf'):
            safe_pdf = (safe_pdf or 'jd') + '.pdf'
        jd_pdf_path = os.path.join(JD_OUTPUT_DIR, safe_pdf)
        with open(jd_pdf_path, 'wb') as pdf_out:
            pdf_out.write(raw_bytes)
        with open(os.path.join(JD_OUTPUT_DIR, 'JD_last_upload.pdf'), 'wb') as last_up:
            last_up.write(raw_bytes)
    except Exception as e:
        print(f'[Server] JD: could not save PDF copy to output/JD: {e}')

    print(f'[Server] /api/jd/upload saved: {JD_CURRENT_FILE}')

    return jsonify({
        'success': True,
        'jd_data': jd_data,
        'rawText': raw_text,
        'saved': True,
        'filename': 'JD_current.json',
    })


@app.route('/api/jd/load', methods=['GET'])
def jd_load_current():
    """Tải JD hiện tại (chỉ một file JD_current.json)."""
    if not os.path.isfile(JD_CURRENT_FILE):
        return jsonify({'success': False, 'message': 'Chua co JD'}), 404
    try:
        with open(JD_CURRENT_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    if not data.get('jd_data'):
        return jsonify({'success': False, 'message': 'Chua co JD'}), 404
    return jsonify({
        'success': True,
        'jd_data': data['jd_data'],
        'rawText': data.get('rawText') or '',
    })


@app.route('/api/jd/text', methods=['POST'])
def upload_jd_text():
    """Nhan JD text truc tiep (khong can PDF) -> Claude extract -> tra ve JSON."""
    data = request.get_json()
    raw_text = (data or {}).get('text', '').strip()
    if not raw_text or len(raw_text) < 20:
        return jsonify({'success': False, 'message': 'Noi dung JD qua ngan'}), 400

    api_key = get_anthropic_key()
    if not api_key:
        return jsonify({'success': False, 'message': 'Thieu ANTHROPIC_API_KEY'}), 503

    if not _quick_doc_check(raw_text, 'jd', api_key):
        return jsonify({
            'success': False,
            'message': 'Nội dung bạn nhập không phải là JD. Vui lòng nhập đúng nội dung mô tả công việc.',
            'errorType': 'wrong_document'
        }), 400

    try:
        jd_data = call_claude_extract_jd(raw_text, api_key)
    except Exception as e:
        print('[Server] /api/jd/text error:', e)
        return jsonify({'success': False, 'message': str(e)}), 500

    payload = {
        'filename': 'JD_current.json',
        'jd_data': jd_data,
        'rawText': raw_text,
        '_meta': {
            'extractedAt': datetime.now().isoformat(),
            'rawTextLength': len(raw_text),
            'source': 'text_input',
        },
    }
    with open(JD_CURRENT_FILE, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f'[Server] /api/jd/text saved: {JD_CURRENT_FILE}')

    return jsonify({
        'success': True,
        'jd_data': jd_data,
        'saved': True,
        'filename': 'JD_current.json',
    })


# ========================================
# HR: danh sách ứng viên (một pool) + quyết định Đậu/Rớt
# ========================================
@app.route('/api/hr/pool', methods=['GET'])
def hr_pool_get():
    """Trả về ứng viên sắp xếp điểm giảm dần.
    Query ?clear=1: xóa toàn bộ ứng viên trong pool rồi mới trả (dùng khi load lại trang — ổn định hơn POST riêng).
    """
    pool = _hr_load_pool()
    if request.args.get('clear') == '1':
        pool['candidates'] = []
        _hr_save_pool(pool)
        pool = _hr_load_pool()
        print('[Server] GET /api/hr/pool?clear=1 — đã làm sạch danh sách ứng viên')
    cands = pool.get('candidates') or []
    try:
        cands = sorted(cands, key=lambda x: float(x.get('score') or 0), reverse=True)
    except Exception:
        pass
    pool['candidates'] = cands
    return jsonify({'success': True, 'pool': pool})


@app.route('/api/hr/pool/clear', methods=['GET', 'POST'])
def hr_pool_clear():
    """Xóa danh sách ứng viên HR (backup nếu client không dùng GET /pool?clear=1)."""
    pool = _hr_load_pool()
    pool['candidates'] = []
    _hr_save_pool(pool)
    print('[Server] /api/hr/pool/clear — đã làm sạch danh sách ứng viên')
    return jsonify({'success': True})


@app.route('/api/hr/candidate/<cid>', methods=['GET'])
def hr_candidate_get(cid):
    """Ứng viên xem thông báo — tra cứu theo id (session pool)."""
    if not cid or len(cid) > 400:
        return jsonify({'success': False, 'message': 'invalid id'}), 400
    pool = _hr_load_pool()
    for c in pool.get('candidates') or []:
        if c.get('id') == cid:
            return jsonify({'success': True, 'candidate': c})
    return jsonify({'success': False, 'message': 'Khong tim thay'}), 404


@app.route('/api/hr/pool/submit', methods=['POST'])
def hr_pool_submit():
    """Ứng viên hoàn tất chấm điểm — ghi vào pool."""
    data = request.get_json() or {}

    job_title = (data.get('jobTitle') or '').strip()
    cand = data.get('candidate') or {}
    name = (cand.get('name') or '').strip() or 'Chưa rõ tên'
    email = (cand.get('email') or '').strip() or ''
    try:
        score = float(cand.get('score'))
    except (TypeError, ValueError):
        score = 0.0
    recommendation = (cand.get('recommendation') or '').strip()
    cid = (cand.get('id') or '').strip() or str(uuid.uuid4())
    ai_strengths = cand.get('aiStrengths')
    ai_development = cand.get('aiDevelopment')
    if not isinstance(ai_strengths, list):
        ai_strengths = []
    if not isinstance(ai_development, list):
        ai_development = []
    ai_summary = (cand.get('aiAnalysisSummary') or '').strip()
    ai_improvement = (cand.get('aiAnalysisImprovement') or '').strip()

    pool = _hr_load_pool()
    if job_title and not (pool.get('jobTitle') or '').strip():
        pool['jobTitle'] = job_title

    candidates = pool.get('candidates') or []
    existing = next((i for i, x in enumerate(candidates) if x.get('id') == cid), None)
    entry = {
        'id': cid,
        'name': name,
        'email': email,
        'score': round(score, 2),
        'recommendation': recommendation,
        'aiStrengths': ai_strengths,
        'aiDevelopment': ai_development,
        'aiAnalysisSummary': ai_summary,
        'aiAnalysisImprovement': ai_improvement or None,
        'analysisJsonFile': None,
        'submittedAt': datetime.now().isoformat(),
        'hrDecision': None,
        'hrFeedback': None,
        'decidedAt': None,
        'notifiedAt': None,
    }
    if existing is not None:
        old = candidates[existing]
        entry['hrDecision'] = old.get('hrDecision')
        entry['hrFeedback'] = old.get('hrFeedback')
        entry['decidedAt'] = old.get('decidedAt')
        entry['notifiedAt'] = old.get('notifiedAt')
        if not ai_strengths and not ai_development:
            entry['aiStrengths'] = old.get('aiStrengths') or []
            entry['aiDevelopment'] = old.get('aiDevelopment') or []
        if not ai_summary:
            entry['aiAnalysisSummary'] = old.get('aiAnalysisSummary') or ''
        if not ai_improvement:
            entry['aiAnalysisImprovement'] = old.get('aiAnalysisImprovement')
        candidates[existing] = entry
    else:
        candidates.append(entry)

    rel_analysis = _hr_save_analysis_snapshot(name, cid, job_title, entry)
    if rel_analysis:
        entry['analysisJsonFile'] = rel_analysis
        if existing is not None:
            candidates[existing] = entry
        else:
            candidates[-1] = entry

    pool['candidates'] = candidates
    _hr_save_pool(pool)
    print(f'[Server] /api/hr/pool/submit candidate={cid} score={score}')
    return jsonify({'success': True, 'candidateId': cid})


@app.route('/api/hr/decision', methods=['POST'])
def hr_decision():
    """HR gửi feedback + Đậu/Rớt; đánh dấu đã thông báo (demo)."""
    data = request.get_json() or {}
    cand_id = (data.get('candidateId') or '').strip()
    decision = (data.get('decision') or '').strip().lower()
    feedback = (data.get('feedback') or '').strip()

    if not cand_id:
        return jsonify({'success': False, 'message': 'Thieu candidateId'}), 400
    if decision not in ('pass', 'fail'):
        return jsonify({'success': False, 'message': 'decision phai la pass hoac fail'}), 400

    pool = _hr_load_pool()
    candidates = pool.get('candidates') or []
    found = None
    for i, x in enumerate(candidates):
        if x.get('id') == cand_id:
            found = i
            break
    if found is None:
        return jsonify({'success': False, 'message': 'Khong tim thay ung vien'}), 404

    candidates[found]['hrDecision'] = decision
    candidates[found]['hrFeedback'] = feedback
    candidates[found]['decidedAt'] = datetime.now().isoformat()
    candidates[found]['notifiedAt'] = datetime.now().isoformat()
    pool['candidates'] = candidates
    _hr_save_pool(pool)

    label = 'Đậu' if decision == 'pass' else 'Rớt'
    print(f'[Server] /api/hr/decision candidate={cand_id} -> {label}')
    return jsonify({
        'success': True,
        'message': f'Da ghi nhan: {label}',
        'candidate': candidates[found],
    })


if __name__ == '__main__':
    print('=' * 50)
    print('CV Analysis Server - http://localhost:5001')
    print(f'JSON output folder: {OUTPUT_DIR} (CV/, JD/, Survey/, hr_pool/)')
    print(f'Model: {MODEL_DIR} | Preprocess: {PREPROCESS_DIR}')
    print('=' * 50)
    print('Routes:')
    print('  POST /api/cv/save        - Luu CV JSON')
    print('  POST /api/cv/upload      - Nhan CV upload')
    print('  POST /api/jd/upload      - Nhan JD upload (PDF)')
    print('  POST /api/jd/text        - Nhan JD text truc tiep')
    print('  GET  /api/jd/load         - Tai JD hien tai (JD_current.json)')
    print('  GET  /api/hr/pool         - Danh sach ung vien (?clear=1 xoa khi load trang)')
    print('  GET/POST /api/hr/pool/clear - Xoa danh sach ung vien')
    print('  GET  /api/hr/candidate/<id>- Chi tiet ung vien (thong bao HR)')
    print('  POST /api/hr/pool/submit - Ung vien nop sau cham diem')
    print('  POST /api/hr/decision    - HR Dau / Rot + feedback')
    print('  POST /api/predict        - Du doan XGBoost')
    print('  POST /api/preprocess     - TF-IDF vector')
    print('  POST /api/score          - Preprocess + XGBoost + SHAP (all-in-one)')
    print('  GET  /api/features       - Lay feature names')
    print('  GET  /api/analysis/<id>  - Lay ket qua phan tich')
    print('=' * 50)
    app.run(port=5001, debug=False)
